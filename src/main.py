import os
import tabula
import itertools
import pandas as pd
from tqdm import tqdm
from materia import Materia
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_df(DF_PATH_OUTPUT: str) -> pd.DataFrame:
    PDF_PATH = input('| PDF Path'.ljust(20) + ' | ')
    print(' '+'-'*20)
    if not os.path.exists('data'): os.mkdir('data')
    if not os.path.exists(DF_PATH_OUTPUT):
        pdf_table = tabula.read_pdf(
            PDF_PATH, 
            pages="all", 
            lattice=True, 
            multiple_tables=True
        )
        df  = pd.concat(pdf_table)
        df['Materia' ] = df['Materia' ].replace('\r', ' ', regex=True)
        df['Profesor'] = df['Profesor'].replace('\r', ' ', regex=True)
        df.to_csv(DF_PATH_OUTPUT, index=False)
    else:
        df = pd.read_csv(DF_PATH_OUTPUT)
    return df

def range_to_intervals(hours: tuple) -> list[str]:
    hour_it     = hours[0]
    hour_ranges = []
    while True:
        start    = str(hour_it   ).zfill(4)
        finish   = str(hour_it+59).zfill(4)
        hour_it += 100
        hour_ranges.append(f'{start}-{finish}')
        if hour_it-41 == hours[1]: break
    return hour_ranges

def get_NRCs(df: pd.DataFrame, classes_to_take: list[str]) -> tuple[dict[int, Materia], list[tuple[str, str]]]:
    NRCs    = {}
    Imparte = []

    for _, row in df.iterrows():
        if row['Materia'] not in classes_to_take: continue
        start, finish = row['Hora'].split('-')
        hours = range_to_intervals((int(start), int(finish)))
        nrc = int(row['NRC'])
        if nrc in NRCs:
            NRCs[nrc].add_hora(hours)
            NRCs[nrc].add_dia(row['Días'])
        else:
            row_dict         = row.to_dict()
            row_dict['Hora'] = hours
            NRCs[nrc] = Materia(row_dict)
        if (nrc, row['Profesor']) not in Imparte:
            Imparte.append((nrc, row['Profesor']))

    return NRCs, Imparte

def group_by_name(NRCs: dict[int, Materia]) -> dict[str, list[str]]:
    classes_by_name = {}
    print('Agrupando clases por nombre...')
    for _, class_ in tqdm(NRCs.items()):
        if class_.MATERIA not in classes_by_name:
            classes_by_name[class_.MATERIA] = [class_.NRC]
        else:
            classes_by_name[class_.MATERIA].append(class_.NRC)
    return classes_by_name

def get_schedules(NRCs: dict[int, Materia], classes_by_name: dict[str, list[str]], prof_blacklist: list[str], hour_ranges: list[str]) -> list[tuple[str]]:
    OUTPUT_PATH        = 'schedules/combinations.txt'
    if not os.path.exists('schedules'): os.mkdir('schedules')
    schedules          = []
    possible_schedules = itertools.product(*classes_by_name.values())

    print('Generando horarios...')
    with open(OUTPUT_PATH, 'w') as file:
        for schedule in tqdm(possible_schedules):
            time_conflicts  = {}
            conflict_exists = False
            for nrc in schedule:
                # Check if there is a professor conflict
                if NRCs[nrc].PROFESOR in prof_blacklist: conflict_exists = True; break
                # Check if there is a hour conflict
                all_hours_in_range = all(hour_range in hour_ranges for hour_range in NRCs[nrc].HORAS)
                if not all_hours_in_range: conflict_exists = True; break
                # Check if there is a schedule conflict
                time_slot = f'{NRCs[nrc].DIAS, NRCs[nrc].HORAS}'
                time_conflicts[time_slot] = time_conflicts.get(time_slot, 0) + 1
                if time_conflicts[time_slot] > 1: conflict_exists = True; break
            if not conflict_exists: 
                file.write(f'{schedule}\n')
                schedules.append(schedule)

    if not schedules: print('No hay horarios disponibles'); exit()
    return schedules

def custom_agg(x: pd.Series) -> str:
    st = set(x.dropna())
    st_str = ', '.join(str(e) for e in st)
    return st_str if st_str else ''

def get_params() -> tuple[list[str], list[str], tuple[int]]:
    classes_to_take = []
    print(' '+'-'*20)
    classes_quantity = int(input('| No. Clases'.ljust(20) + ' | '))
    print(' '+'-'*20)
    for i in range(classes_quantity):
        classes_to_take.append(input(f'| Clase {i+1}'.ljust(20) + ' | '))
        print(' '+'-'*20)
        # classes_to_take.append(input(f'Nombre de la materia {i+1}: '.ljust(25)))
    prof_blacklist = input('| Profes a Evitar'.ljust(20) + ' | ').lower().split(', ')
    print(' '+'-'*20)
    start_hour = int(input('| Desde: '.ljust(20) + ' | '))
    print(' '+'-'*20)
    end_hour   = int(input('| Hasta: '.ljust(20) + ' | '))
    print(' '+'-'*20)
    hour_range = (start_hour*100, end_hour*100-41)
    return classes_to_take, prof_blacklist, hour_range

def format_xlsx(PATH: str, color_map: dict[str, int], NRCs: dict[int, Materia]) -> None:
    book = load_workbook(f'{PATH}.xlsx')
    sheet = book.active

    colors = [
        'FFB0E0E6',  # Pastel Blue
        'FFF49AC2',  # Pastel Magenta
        'FF77DD77',  # Pastel Green
        'FFAEC6CF',  # Pastel Pink
        'FFFFB347',  # Pastel Orange
        'FFFDFD96',  # Pastel Yellow
        'FFFF6961',  # Pastel Red
        'FFCB99C9',  # Pastel Purple
        'FFB39EB5',  # Pastel Violet
        'FFCC99FF'   # Pastel Lavender
    ]

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    print('Aplicando formato al archivo...')
    for name, i in tqdm(color_map.items()):
        pattern = PatternFill(start_color=colors[i], end_color=colors[i], fill_type='solid')
        #  max_col=sheet.max_column, max_row=sheet.max_row):
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=8):
            for cell in row:
                try:
                    if NRCs[int(cell.value)].MATERIA == name:
                        cell.fill = pattern
                except:
                    pass
                    

    book.save(f'{PATH}.xlsx')
    print(f'Horarios guardados en {PATH}.xlsx')

def get_priority(schedule_df: pd.DataFrame, NRCs: dict[int, Materia], white_list: list[str]) -> int:
    df       = schedule_df.copy()
    droped   = df[(df['Martes'] == '') & (df['Miércoles'] == '')]
    aux_df   = df[(df['Martes'] != '') | (df['Miércoles'] != '')]

    start        = int(aux_df.index[0].split('-')[0])
    end          = int(aux_df.index[-1].split('-')[1])
    active_hours = range_to_intervals((start, end))

    priority   = 0
    dead_hours = len(set(active_hours) & set(droped.index))

    for i, (j, nrc) in enumerate(aux_df['Lunes'].items()):
        if nrc == '' and i == 0: continue
        if nrc == '' and i == len(aux_df)-1: continue
        if nrc == '' and j == active_hours[-2]: priority += 5; continue
        if nrc == '' and j == active_hours[1] : priority += 5; continue
        if nrc == '': dead_hours+=1; continue
        if NRCs[int(nrc)].PROFESOR in white_list: priority -= 10
    
    priority += dead_hours * 20
    return priority

def group_by_priority(
            schedules: list[tuple[str]], 
            NRCs: dict[int, Materia], 
            hour_intervals: list[str], 
            white_list: list[str]
    ) -> tuple[dict[int, list[tuple[pd.DataFrame, pd.DataFrame]]], dict[str, int]]:
    
    schedules_by_priority = {}
    color_map = {}
    i = 0
    print('Obteniendo los mejores horarios...')
    for schedule in tqdm(schedules):
        schedule_df = pd.DataFrame(columns=['HORAS', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes'])
        classes_df_keys = ['NRC', 'MATERIA', 'HORAS', 'PROFESOR', 'SALON']
        classes_df = pd.DataFrame(columns=classes_df_keys)
        for nrc in schedule:
            schedule_df = pd.concat([schedule_df, pd.DataFrame({
                'HORAS'    : [NRCs[nrc].HORAS[0]], 
                'Lunes'    : nrc if 'L' in NRCs[nrc].DIAS and 'M' in NRCs[nrc].DIAS else None,
                'Martes'   : nrc if 'A' in NRCs[nrc].DIAS else None,
                'Miércoles': nrc if 'M' in NRCs[nrc].DIAS else None,
                'Jueves'   : nrc if 'J' in NRCs[nrc].DIAS else None,
                'Viernes'  : nrc if 'V' in NRCs[nrc].DIAS else None
            })])
            schedule_df = pd.concat([schedule_df, pd.DataFrame({
                'HORAS'    : [NRCs[nrc].HORAS[1]], 
                'Lunes'    : nrc if 'L' in NRCs[nrc].DIAS and 'A' in NRCs[nrc].DIAS else None,
                'Martes'   : nrc if 'A' in NRCs[nrc].DIAS else None,
                'Miércoles': nrc if 'M' in NRCs[nrc].DIAS else None,
                'Jueves'   : nrc if 'J' in NRCs[nrc].DIAS else None,
                'Viernes'  : nrc if 'V' in NRCs[nrc].DIAS else None
            })])

            class_temp = pd.DataFrame({key: NRCs[nrc].__dict__[key] for key in classes_df_keys}, index=[0, 1])
            classes_df = pd.concat([classes_df, class_temp])

            if NRCs[nrc].MATERIA not in color_map: color_map[NRCs[nrc].MATERIA] = i; i+=1

        for hour in hour_intervals:
            if hour not in schedule_df['HORAS'].values:
                schedule_df = pd.concat([schedule_df, pd.DataFrame({
                    'HORAS'    : [hour], 
                    'Lunes'    : None,
                    'Martes'   : None,
                    'Miércoles': None,
                    'Jueves'   : None,
                    'Viernes'  : None
                })])

        grouped_df = schedule_df.groupby(['HORAS']).agg(custom_agg)
        grouped_df = grouped_df.astype(str)
        classes_df = classes_df.astype(str)
        classes_df['PROFESOR'] = classes_df['PROFESOR'].str.title()
        classes_df = classes_df.sort_values(by=['HORAS'])
        classes_df = classes_df[['NRC', 'MATERIA', 'PROFESOR', 'SALON']]
        classes_df = classes_df.drop_duplicates()

        priority = get_priority(grouped_df, NRCs, white_list)

        if priority not in schedules_by_priority: schedules_by_priority[priority] = [(grouped_df, classes_df)]
        else: schedules_by_priority[priority].append((grouped_df, classes_df))

    return dict(sorted(schedules_by_priority.items())), color_map
    # return schedules_by_priority, color_map

def save_schedules(
            SCHEDULES_PATH: str, 
            schedules_by_priority: dict[int, list[tuple[pd.DataFrame, pd.DataFrame]]]
    ) -> None:

    if not os.path.exists(SCHEDULES_PATH.split('/')[0]):
        os.mkdir(SCHEDULES_PATH.split('/')[0])
    
    print('Guardando horarios...')
    with pd.ExcelWriter(f'{SCHEDULES_PATH}.xlsx') as writer:
        row_count = 0
        for _, schedules in tqdm(schedules_by_priority.items()):
            for schedule in schedules:
                grouped_df, classes_df = schedule
                grouped_df.to_excel(
                    writer, 
                    sheet_name='Sheet1', 
                    startrow=row_count
                )
                classes_df.to_excel(
                    writer, 
                    sheet_name='Sheet1', 
                    startrow=row_count, 
                    startcol=len(grouped_df.columns) + 2, 
                    index=False
                )
                row_count += len(grouped_df) + 4

def without_nrc(schedules: list[tuple[str]], NRCs_blacklist: list[str]) -> list[tuple[str]]:
    schedules_without_nrc = []
    for schedule in schedules:
        if not any(str(nrc) in NRCs_blacklist for nrc in schedule):
            schedules_without_nrc.append(schedule)
    return schedules_without_nrc

def main():
    SCHEDULES_PATH = 'schedules/schedules'
    DF_PATH_OUTPUT = 'data/classes.csv'

    classes_to_take, prof_blacklist, hour_range = get_params()
    df              = get_df(DF_PATH_OUTPUT)
    NRCs, Imparte   = get_NRCs(df, classes_to_take)
    classes_by_name = group_by_name(NRCs)
    hour_intervals  = range_to_intervals(hour_range)
    schedules       = get_schedules(NRCs, classes_by_name, prof_blacklist, hour_intervals)
    schedules_by_priority, color_map = group_by_priority(schedules, NRCs, hour_intervals, [])
    save_schedules(SCHEDULES_PATH, schedules_by_priority)
    format_xlsx(SCHEDULES_PATH, color_map, NRCs)

    n = 1
    while input('¿Desea eliminar horarios con NRCs específicos? (y/n): ').lower() == 'y':
        NRCs_blacklist = input('NRCs a eliminar: ').split(', ')
        schedules_aux = without_nrc(schedules, NRCs_blacklist)
        schedules_by_priority_aux, color_map = group_by_priority(schedules_aux, NRCs, hour_intervals, [])
        save_schedules(f'{SCHEDULES_PATH}{n}', schedules_by_priority_aux)
        format_xlsx(f'{SCHEDULES_PATH}{n}', color_map, NRCs)
        n += 1

if __name__ == '__main__':
    main()
